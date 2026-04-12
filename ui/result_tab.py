"""
Author              : MARCUS THAM WAI KEAN
DATE CREATED        : 3-MARCH-2026
DATE COMPLETED      : 31-MARCH-2026
ROLL-OUT VERSION    : v1.0.0
DESCRIPTION         : This is the third tab on the main window that display all the final results from the iteration
                      computed from CalculationTab().
"""

from PyQt6.QtWidgets import (
    QWidget,QLabel,QFileDialog,QPushButton,QTableWidgetItem,QVBoxLayout,QFrame,QTableWidget,QHeaderView,
      QAbstractItemView, QMessageBox
      )

from PyQt6.QtCore import Qt

#══════════════════════════════════════════════════════════════════════════════════════════════════════════════════════
#─────CLASS 5.0────────────────────────────────────────────────────────────────────────────────────────────────────────
"""
Create a class named finalResultTab() to display the final iteration results computed in module_number_calculation.py.
"""
class finalResultTab(QWidget):

    #─────5.1: INITIALISE THE CLASS WITH PRE-DEFINED OBJECTS AND METHODS───────────────────────────────────────────────
    def __init__(self):
        super().__init__()

        self.build_ui()
        
    #─────5.2: HELPER METHOD #0────────────────────────────────────────────────────────────────────────────────────────
    def build_ui(self):
        
        #Set up vertical box layout to place widgets
        root = QVBoxLayout(self)
        root.setContentsMargins(24,24,24,24)
        root.setSpacing(10)

        #Create label widgets in the vertical layout for title
        title = QLabel("Iteration Results")
        title.setObjectName("title")
        root.addWidget(title)
        
        #Create frame widgets in the vertical layout for divider line
        divider = QFrame()
        divider.setFrameShape(QFrame.Shape.HLine)
        divider.setObjectName("divider")
        root.addWidget(divider)

        #Create a table widget that will display the table populated by the runCalculator() method in CalculationTab()
        self.final_result_table = QTableWidget()
        self.final_result_table.setObjectName("table")

        #Set the horizontal header column to expand to fill the entire table automatically
        self.final_result_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        #Set edit trigger rule such that edition on any table cells is restricted
        self.final_result_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        root.addWidget(self.final_result_table)
        
        #Add an export button that connects to HELPER METHOD #2 to export result table to excel file
        self.export_button= QPushButton("Export to Excel")
        self.export_button.setObjectName("primaryBtn")
        self.export_button.clicked.connect(self.export_to_excel)
        root.addWidget(self.export_button)

    #─────5.3: HELPER METHOD #1────────────────────────────────────────────────────────────────────────────────────────
    """
    This helper method takes final iterated resuls in form of dictionary and generate a final result table on the 
    table widget.
    """
    def populate_result_table(self,final_iterated_results:dict):

        #From the result dictionary, set the row number for the table widget
        self.final_result_table.setRowCount(len(final_iterated_results))

        #From the result dictionary, set the column count for the table widget which is a fixed integer of 4
        self.final_result_table.setColumnCount(4)

        #Label the column header
        self.final_result_table.setHorizontalHeaderLabels([
            "Parameters", "Unit", "1st Pass", "2nd Pass"
        ])

        """
        Populate table result with 'for' loop:
            - The dict.items() method will return each key and values pair in the form of tuple
            - enumerate() gives each key and values pair an index for looping, where i is to loop each index
            - The '_' means the key component will not be used in the loop
        """
        for i, (_,value) in enumerate(final_iterated_results.items()):
            
            """
            From module_number_calculation.py, in the final iterated result dictionary, each dictionary key has
            4 values. Create 4 new objects to place each value from the dictionary where:
                1. param = parameters column
                2. unit = unit column
                3. val_1p = 1st pass result value
                4. val_2p = 2nd pass result value
            """
            param, unit, val_1p, val_2p = value

            """
            1. QTableWidgetItem() represents a single cell on the QTableWIdget()

            2. Set text alignment as center

            3. Finally, set the table widget item to the QTableWidget
            """
            item_param = QTableWidgetItem(param)
            item_param.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.final_result_table.setItem(i, 0, item_param)        #For each row, set the param value to 1st column
            
            item_unit = QTableWidgetItem(unit)
            item_unit.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.final_result_table.setItem(i, 1, item_unit)        #For each row, set the unit value to 2nd column
            
            #Set conditions where if result is empty, display '-' instead of leaving it blank, else, convert to string
            display_1p = "-" if val_1p is None else str(val_1p) 
            display_2p = "-" if val_2p is None else str(val_2p)

            item_1p = QTableWidgetItem(display_1p)
            item_1p.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.final_result_table.setItem(i, 2, item_1p)         #For each row, set the 1P result value to 2nd column

            item_2p = QTableWidgetItem(display_2p)
            item_2p.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.final_result_table.setItem(i, 3, item_2p)         #For each row, set the 2P result value to 2nd column

    #─────5.4: HELPER METHOD #2────────────────────────────────────────────────────────────────────────────────────────
    """
    This helper method allows user to export the result table into a excel file with naming of own choice to desired
    directory.
    """
    def export_to_excel(self):

        """
        1. QFileDialog is one widget from QtWidgets that open a file browser that shows files and folder from system
        2. The '.getSaveFileName() returns a tuple of (filepath,selected_filter) where selected_filter is ignored here
        """
        filepath,_ = QFileDialog.getSaveFileName(
            self,
            "Export Results",               #This will be the name of the file dialog
            "iteration_results.xlsx",       #This will be the file name to be saved
            "Excel Files (*.xlsx)"          #This will be the file type filter
        ) 

        #If user did not create a filepath to save or close the dialog, display nothing instead of error pop-up
        if not filepath:
            return  
        
    
        try:
            import openpyxl as xl
            from openpyxl.styles import Font,Alignment,PatternFill

            #Create an empty workbook by calling openxl
            workbook = xl.Workbook()

            #Get the default active worksheet
            worksheet = workbook.active

            #Set the worksheet naming as desired naming
            worksheet.title = "Iteration Results"

            #Set up namings for each column header
            headers = ["Parameters", "Unit", "1st Pass", "2nd Pass"]

            #Set the styles for the column headers
            header_fill = PatternFill(start_color="1B6B3A", end_color="1B6B3A", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            #Loop through each column header to fill the header styles and information
            for col,header in enumerate(headers,start=1):

                cell = worksheet.cell(row = 1, column=col, value = header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal = "center",vertical= "center")

            #Loop through each row in the final result table
            for row_idx in range(self.final_result_table.rowCount()):
               
                #The data row starts from row 2 as row 1 is the header row
                excel_row = row_idx + 2

                #For each column in the final result table
                for col_idx in range(self.final_result_table.columnCount()):
                    

                    item = self.final_result_table.item(row_idx, col_idx)
                    
                    #Extract the value from each row and column pair from the final result table
                    value = item.text() if item else ""

                    #Input the extracted value to matching row and column pairs
                    cell = worksheet.cell(row=excel_row, column=col_idx + 1, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            #Set up custom width for each column where the alphabets is the default column index in excel
            worksheet.column_dimensions["A"].width = 35
            worksheet.column_dimensions["B"].width = 12
            worksheet.column_dimensions["C"].width = 18
            worksheet.column_dimensions["D"].width = 18

            #Iterate through the worksheet using .iter_rows()
            for row in worksheet.iter_rows():

                #Set up the height of each row 
                worksheet.row_dimensions[row[0].row].height = 20

            #Finally, save the workbook to a user defined filepath
            workbook.save(filepath)

            #Pop up a message box to tell user about successful result table export
            QMessageBox.information(self, "Export Successful!")

        #Error pop up if any errors occur
        except Exception as e:
            QMessageBox.critical(self, "Export failed, please try again!")